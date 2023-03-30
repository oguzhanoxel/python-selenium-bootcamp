import pytest
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.action_chains import ActionChains
from pathlib import Path
from datetime import date
from time import sleep
import openpyxl
from constants import global_constants


class Demo:

	def setup_method(self):
		self.driver = webdriver.Chrome()
		self.driver.maximize_window()
		self.driver.get(global_constants.URL)
		self.today = str(date.today())
		Path(self.today).mkdir(exist_ok=True)

	def teardown_method(self):
		self.driver.quit()

	def test_demo(self):
		text = "Hello"
		assert text == "Hello"

	def getData():
		excelFile = openpyxl.load_workbook("data/invalid_login.xlsx")
		selectedSheet = excelFile["Sayfa1"]
		
		totalRows = selectedSheet.max_row
		data=[]
		for i in range(2, totalRows+1):
			username = selectedSheet.cell(i, 1).value
			password = selectedSheet.cell(i, 2).value
			tupleData = (username, password)
			data.append(tupleData)

		return data

	# @pytest.mark.skip()
	@pytest.mark.parametrize("username,password", getData())
	def test_invalid_login(self, username, password):
		self.waitForElementVisible((By.ID, "user-name"))
		self.waitForElementVisible((By.ID, "password"))

		usernameInput = self.driver.find_element(By.ID, "user-name")
		passwordInput = self.driver.find_element(By.ID, "password")
		sleep(1)

		usernameInput.send_keys(username)
		passwordInput.send_keys(password)

		loginButton = self.driver.find_element(By.ID, "login-button")
		sleep(1)
		loginButton.click()
		sleep(1)
		errorMessage = self.driver.find_element(By.XPATH, "/html/body/div/div/div[2]/div[1]/div/div/form/div[3]/h3")

		self.driver.save_screenshot(f"{self.today}/test-invalid-login-{username}-{password}.png")
		assert errorMessage.text == "Epic sadface: Username and password do not match any user in this service"

	def waitForElementVisible(self, locator, timeout=5):
		WebDriverWait(self.driver, timeout).until(
			expected_conditions
			.visibility_of_element_located(locator))

